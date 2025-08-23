"""
Standard Excel Export Module

Creates organized Excel files with one file per model containing multiple sheets
for different sample counts. This is the main results exporter.
"""

import pandas as pd
import numpy as np
from pathlib import Path
from typing import List, Dict, Any
from datetime import datetime

try:
    import openpyxl
    from openpyxl.styles import PatternFill, Font
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False

import sys
from pathlib import Path

sys.path.append(str(Path(__file__).parent.parent))

from utils.data_structures import ParsedResult


class ExcelExporter:
    """
    Standard Excel exporter for scaling results.
    
    Creates one Excel file per model with multiple sheets (one per sample count).
    Each sheet contains results from different seeds for easy averaging.
    """
    
    def __init__(self):
        self.standard_columns = [
            'seed', 'timestamp', 'accuracy', 'avg_voting_confidence', 
            'avg_time_per_question', 'avg_tokens_per_second',
            'avg_ttft', 'avg_decode_time', 'avg_input_tokens', 'avg_output_tokens',
            'total_questions', 'correct_answers', 'total_samples_generated'
        ]
    
    def export_by_model_multisheet(self, results: List[ParsedResult], output_dir: str = "./parsed_results") -> None:
        """
        Export results with one Excel file per model, multiple sheets per sample count.
        
        Structure: 
        parsed_results/
        ├── DeepSeek-R1-Distill-Qwen-1.5B.xlsx (sheets: 1_samples, 2_samples, 4_samples, etc.)
        ├── DeepSeek-R1-Distill-Qwen-14B.xlsx
        └── scaling_summary.xlsx
        """
        if not EXCEL_AVAILABLE:
            print("❌ Excel export not available. Install openpyxl: pip install openpyxl")
            return
            
        output_path = Path(output_dir)
        output_path.mkdir(exist_ok=True)
        
        # Convert results to DataFrame
        df = self._create_results_dataframe(results)
        
        if df.empty:
            print("❌ No data to export")
            return
        
        print(f"📊 Starting Excel Export (Multi-Sheet Format)")
        print(f"📊 Total results: {len(df)}")
        
        # Export one file per model with multiple sheets
        for model_name in df['model_name'].unique():
            model_df = df[df['model_name'] == model_name].copy()
            self._create_model_excel_file(model_df, model_name, output_path)
        
        # Create overall scaling summary
        self._create_scaling_summary(df, output_path)
        
        print(f"\n✅ Excel export complete! Files saved in: {output_path}")
    
    def _create_results_dataframe(self, results: List[ParsedResult]) -> pd.DataFrame:
        """Convert ParsedResult objects to DataFrame."""
        rows = []
        
        # Time columns that need conversion from ms to seconds
        time_columns_ms = ['avg_time_per_question', 'avg_ttft', 'avg_decode_time']
        
        for result in results:
            row = {
                'model_name': result.metadata.model_name,
                'seed': result.metadata.seed,
                'num_samples': result.metadata.num_samples,
                'timestamp': result.metadata.timestamp,
                'accuracy': result.metrics.accuracy,
                'avg_voting_confidence': result.metrics.avg_voting_confidence,
                'avg_time_per_question': result.metrics.avg_time_per_question,
                'avg_tokens_per_second': result.metrics.avg_tokens_per_second,
                'avg_ttft': result.metrics.avg_ttft,
                'avg_decode_time': result.metrics.avg_decode_time,
                'avg_input_tokens': result.metrics.avg_input_tokens,
                'avg_output_tokens': result.metrics.avg_output_tokens,
                'total_questions': result.metrics.total_questions,
                'correct_answers': result.metrics.correct_answers,
                'total_samples_generated': result.metrics.total_samples_generated,
                'avg_power_consumption': result.metrics.avg_power_consumption,
                'total_energy_consumed': result.metrics.total_energy_consumed,
            }
            
            # Add system stats if available
            if result.system_stats:
                row.update({
                    'avg_ram_percent': result.system_stats.get('avg_ram_percent'),
                    'avg_cpu_percent': result.system_stats.get('avg_cpu_percent'),
                    'avg_gpu_percent': result.system_stats.get('avg_gpu_percent'),
                })
            
            rows.append(row)
        
        df = pd.DataFrame(rows)
        
        # Convert time columns from milliseconds to seconds
        for col in time_columns_ms:
            if col in df.columns:
                df[col] = df[col] / 1000.0  # Convert ms to seconds
        
        # Round all numeric columns to 2 decimal places
        numeric_columns = df.select_dtypes(include=[np.number]).columns
        for col in numeric_columns:
            if col not in ['seed', 'num_samples', 'total_questions', 'correct_answers', 'total_samples_generated']:
                df[col] = df[col].round(2)
        
        # Remove columns that are entirely empty/null
        df = df.dropna(axis=1, how='all')
        
        # Remove columns that have only empty strings or zeros
        for col in df.columns:
            if df[col].dtype == 'object':  # String columns
                if df[col].fillna('').astype(str).str.strip().eq('').all():
                    df = df.drop(columns=[col])
            elif df[col].dtype in ['float64', 'int64']:  # Numeric columns
                if df[col].fillna(0).eq(0).all():
                    df = df.drop(columns=[col])
        
        return df
    
    def _create_model_excel_file(self, model_df: pd.DataFrame, model_name: str, output_path: Path) -> None:
        """Create one Excel file per model with multiple sheets."""
        excel_filename = f"{model_name}.xlsx"
        excel_path = output_path / excel_filename
        
        print(f"📄 Creating {excel_filename}")
        
        sample_counts = sorted(model_df['num_samples'].unique())
        
        with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
            for sample_count in sample_counts:
                sample_df = model_df[model_df['num_samples'] == sample_count].copy()
                sheet_name = f"{sample_count}_samples"
                
                print(f"  📋 Creating sheet: {sheet_name}")
                
                available_columns = [col for col in self.standard_columns if col in sample_df.columns]
                export_df = sample_df[available_columns].copy()
                
                export_df = export_df.sort_values('seed').reset_index(drop=True)
                
                summary_df = self._create_summary_rows(export_df, available_columns)
                final_df = pd.concat([export_df, summary_df], ignore_index=True)
                
                final_df.to_excel(writer, sheet_name=sheet_name, index=False)
                
                self._format_worksheet(writer.sheets[sheet_name], len(export_df), available_columns)
        
        print(f"    ✅ Created: {excel_path}")
    
    def _create_summary_rows(self, df: pd.DataFrame, columns: List[str]) -> pd.DataFrame:
        """Create summary statistics rows."""
        summary_rows = []
        
        numeric_columns = df.select_dtypes(include=[np.number]).columns
        
        if len(numeric_columns) > 0:
            means = df[numeric_columns].mean()
            stds = df[numeric_columns].std()
            mins = df[numeric_columns].min()
            maxs = df[numeric_columns].max()
            
            mean_row = ['MEAN'] + [means.get(col, '') for col in columns[1:]]
            std_row = ['STD'] + [stds.get(col, '') for col in columns[1:]]
            min_row = ['MIN'] + [mins.get(col, '') for col in columns[1:]]
            max_row = ['MAX'] + [maxs.get(col, '') for col in columns[1:]]
            
            summary_rows = [mean_row, std_row, min_row, max_row]
        
        return pd.DataFrame(summary_rows, columns=columns)
    
    def _format_worksheet(self, worksheet, data_rows: int, columns: List[str]) -> None:
        """Format Excel worksheet with colors and proper column widths."""
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 20)
            worksheet.column_dimensions[column_letter].width = adjusted_width
        
        summary_start_row = data_rows + 2
        
        for row_idx in range(summary_start_row, summary_start_row + 4):
            for col_idx in range(1, len(columns) + 1):
                cell = worksheet.cell(row=row_idx, column=col_idx)
                if row_idx == summary_start_row:  # MEAN row
                    cell.fill = PatternFill(start_color="E8F5E8", end_color="E8F5E8", fill_type="solid")
                    cell.font = Font(bold=True)
                elif row_idx == summary_start_row + 1:  # STD row
                    cell.fill = PatternFill(start_color="FFF8DC", end_color="FFF8DC", fill_type="solid")
                elif row_idx == summary_start_row + 2:  # MIN row
                    cell.fill = PatternFill(start_color="E0F2F1", end_color="E0F2F1", fill_type="solid")
                elif row_idx == summary_start_row + 3:  # MAX row
                    cell.fill = PatternFill(start_color="FCE4EC", end_color="FCE4EC", fill_type="solid")
    
    def _create_scaling_summary(self, df: pd.DataFrame, output_path: Path) -> None:
        """Create overall scaling summary across all models."""
        # Load prefill latency data
        prefill_latencies = self._load_prefill_latencies()
        
        summary_data = []
        
        for model_name in df['model_name'].unique():
            model_df = df[df['model_name'] == model_name]
            
            for sample_count in sorted(model_df['num_samples'].unique()):
                sample_data = model_df[model_df['num_samples'] == sample_count]
                
                # Get prefill latency for this model
                prefill_latency = prefill_latencies.get(model_name, 0.0)
                
                # Calculate corrected decode time (total time - prefill latency)
                avg_total_time = sample_data['avg_time_per_question'].mean()
                avg_decode_time_corrected = max(0, avg_total_time - prefill_latency)  # Ensure non-negative
                
                summary_row = {
                    'model_name': model_name,
                    'num_samples': sample_count,
                    'avg_accuracy': sample_data['accuracy'].mean(),
                    'std_accuracy': sample_data['accuracy'].std(),
                    'max_accuracy': sample_data['accuracy'].max(),
                    'avg_time_per_question': avg_total_time,
                    'avg_tokens_per_second': sample_data['avg_tokens_per_second'].mean(),
                    'avg_ttft': sample_data['avg_ttft'].mean() if 'avg_ttft' in sample_data.columns else None,
                    'avg_decode_time': avg_decode_time_corrected,  # Corrected decode time
                    'prefill_latency': prefill_latency,  # Show prefill latency for reference
                    'num_runs': len(sample_data)
                }
                summary_data.append(summary_row)
        
        summary_df = pd.DataFrame(summary_data)
        summary_path = output_path / "scaling_summary.xlsx"
        
        with pd.ExcelWriter(summary_path, engine='openpyxl') as writer:
            summary_df.to_excel(writer, sheet_name="Scaling_Summary", index=False)
            
            ws = writer.sheets["Scaling_Summary"]
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 18)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Highlight headers
            for col_idx in range(1, len(summary_df.columns) + 1):
                cell = ws.cell(row=1, column=col_idx)
                cell.fill = PatternFill(start_color="B3E5FC", end_color="B3E5FC", fill_type="solid")
                cell.font = Font(bold=True)
        
        print(f"    📊 Created: {summary_path}")
    
    def _load_prefill_latencies(self) -> Dict[str, float]:
        """Load prefill latencies from prefill_latency.csv file."""
        prefill_file = Path("results/tokens/prefill_latency.csv")
        prefill_latencies = {}
        
        if not prefill_file.exists():
            print(f"⚠️  Prefill latency file not found: {prefill_file}")
            print("   Using zero prefill latency for all models")
            return prefill_latencies
        
        try:
            import pandas as pd
            df = pd.read_csv(prefill_file)
            
            for _, row in df.iterrows():
                model_name = row['model']  # e.g., 'DSR1-Qwen-14B'
                prefill_latency = float(row['prefill_latency_seconds'])
                
                # Map short names to full names
                if 'DSR1-Qwen-14B' in model_name:
                    full_name = 'DeepSeek-R1-Distill-Qwen-14B'
                elif 'DSR1-Qwen-1.5B' in model_name:
                    full_name = 'DeepSeek-R1-Distill-Qwen-1.5B'
                elif 'DSR1-LLaMA-8B' in model_name or 'DSR1-Llama-8B' in model_name:
                    full_name = 'DeepSeek-R1-Distill-Llama-8B'
                else:
                    full_name = model_name
                
                prefill_latencies[full_name] = prefill_latency
                print(f"   Loaded prefill latency for {full_name}: {prefill_latency:.3f}s")
            
        except Exception as e:
            print(f"⚠️  Error loading prefill latencies: {e}")
            print("   Using zero prefill latency for all models")
        
        return prefill_latencies